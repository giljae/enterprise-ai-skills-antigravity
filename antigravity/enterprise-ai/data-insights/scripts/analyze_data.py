#!/usr/bin/env python3
"""
Simple data analysis - answers 7 key questions in plain English
"""
from pathlib import Path
import sys

pd = None
np = None

def load_dependencies():
    """Import heavy data dependencies only when analysis actually runs."""
    global pd, np
    if pd is not None and np is not None:
        return
    try:
        import pandas as pandas_module
        import numpy as numpy_module
    except Exception as exc:
        print("Error: data-insights dependencies are not available or are incompatible.", file=sys.stderr)
        print("Install them with: pip install -r scripts/requirements.txt", file=sys.stderr)
        print(f"Details: {exc}", file=sys.stderr)
        sys.exit(1)
    pd = pandas_module
    np = numpy_module

def clean_data(df):
    """Basic data cleaning"""
    # Remove duplicates
    df = df.drop_duplicates()
    
    # Identify potential date columns to skip
    date_cols = [col for col in df.columns if 'date' in col.lower() or 'time' in col.lower()]
    
    # Convert string percentages to numbers
    for col in df.select_dtypes(include=['object']).columns:
        if col in date_cols:
            continue
        if df[col].astype(str).str.contains('%').any():
            df[col] = df[col].str.replace('%', '').astype(float) / 100
    
    # Convert currency strings to numbers
    for col in df.select_dtypes(include=['object']).columns:
        if col in date_cols:
            continue
        if df[col].astype(str).str.contains('$', regex=False).any():
            try:
                df[col] = df[col].str.replace('$', '').str.replace(',', '').astype(float)
            except:
                pass  # Skip if conversion fails
    
    # Strip whitespace
    for col in df.select_dtypes(include=['object']).columns:
        try:
            df[col] = df[col].str.strip()
        except:
            pass
    
    return df

def analyze_data(df):
    """Run all 7 core analyses"""
    results = {
        'insights': [],
        'tables': {}
    }
    
    # Identify numeric and time columns
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    time_cols = [col for col in df.columns if 'date' in col.lower() or 'time' in col.lower() or 'period' in col.lower()]
    
    if not numeric_cols:
        results['insights'].append("⚠️ No numeric data found to analyze")
        return results
    
    # Main metric is first numeric column
    main_metric = numeric_cols[0]
    
    # 1. COMPARE TO TARGET/GOAL
    target_cols = [col for col in df.columns if 'target' in col.lower() or 'goal' in col.lower() or 'budget' in col.lower()]
    if target_cols:
        # Find matching metric for target
        target_col = target_cols[0]
        actual_total = df[main_metric].sum()
        target_total = df[target_col].sum()
        
        if target_total > 0:
            variance_pct = ((actual_total - target_total) / target_total) * 100
            
            if variance_pct >= 5:
                results['insights'].append(f"✅ {main_metric}: Beat target by {variance_pct:.0f}%")
            elif variance_pct <= -5:
                results['insights'].append(f"⚠️ {main_metric}: Missed target by {abs(variance_pct):.0f}%")
            else:
                results['insights'].append(f"➡️ {main_metric}: On target (within 5%)")
            
            # Add summary table
            results['tables']['Target Comparison'] = pd.DataFrame({
                'Metric': [main_metric],
                'Actual': [actual_total],
                'Target': [target_total],
                'Variance %': [variance_pct]
            })
    
    # 2. TREND OVER TIME
    if time_cols and len(df) >= 3:
        time_col = time_cols[0]
        df_sorted = df.sort_values(time_col)
        
        # Calculate simple trend
        first_value = df_sorted[main_metric].iloc[0]
        last_value = df_sorted[main_metric].iloc[-1]
        
        if first_value > 0:
            change_pct = ((last_value - first_value) / first_value) * 100
            
            if change_pct > 10:
                results['insights'].append(f"📈 Trending up: {main_metric} increased {change_pct:.0f}% from start to end")
            elif change_pct < -10:
                results['insights'].append(f"📉 Trending down: {main_metric} decreased {abs(change_pct):.0f}% from start to end")
            else:
                results['insights'].append(f"➡️ Stable: {main_metric} stayed relatively flat (changed {change_pct:.0f}%)")
        
        # Add trend table
        trend_data = df_sorted[[time_col, main_metric]].copy()
        trend_data['Change from Previous'] = trend_data[main_metric].diff()
        trend_data['% Change'] = trend_data[main_metric].pct_change() * 100
        results['tables']['Trend Over Time'] = trend_data
    
    # 3. SPOT UNUSUAL CHANGES
    if time_cols and len(df) >= 3:
        df_sorted = df.sort_values(time_col)
        changes = df_sorted[main_metric].pct_change() * 100
        
        # Flag changes > 30%
        unusual = []
        for idx in range(1, len(changes)):
            if pd.notna(changes.iloc[idx]) and abs(changes.iloc[idx]) > 30:
                period = df_sorted.iloc[idx][time_col]
                value = df_sorted.iloc[idx][main_metric]
                change = changes.iloc[idx]
                
                if change > 0:
                    unusual.append(f"📈 Spike: {main_metric} jumped {change:.0f}% in {period}")
                else:
                    unusual.append(f"📉 Drop: {main_metric} fell {abs(change):.0f}% in {period}")
        
        if unusual:
            results['insights'].extend(unusual[:2])  # Top 2 only
        else:
            results['insights'].append(f"✓ No unusual spikes or drops detected in {main_metric}")
    
    # 4. COMPARE SEGMENTS
    categorical_cols = []
    for col in df.columns:
        if df[col].dtype == 'object' or df[col].dtype.name == 'category':
            unique_count = df[col].nunique()
            # Skip date/time columns
            if 'date' in col.lower() or 'time' in col.lower() or 'week' in col.lower() or 'month' in col.lower():
                continue
            if 2 <= unique_count <= 10:  # Between 2 and 10 unique values
                categorical_cols.append(col)
    
    if categorical_cols:
        segment_col = categorical_cols[0]
        segment_summary = df.groupby(segment_col)[main_metric].agg(['sum', 'mean', 'count']).round(0)
        segment_summary.columns = ['Total', 'Average', 'Count']
        segment_summary = segment_summary.sort_values('Total', ascending=False)
        
        best = segment_summary.index[0]
        worst = segment_summary.index[-1]
        
        results['insights'].append(f"🏆 Best segment: {best} (total: {segment_summary.loc[best, 'Total']:,.0f})")
        results['insights'].append(f"📍 Weakest segment: {worst} (total: {segment_summary.loc[worst, 'Total']:,.0f})")
        
        results['tables']['Segment Comparison'] = segment_summary
    
    # 5. TOP/BOTTOM PERFORMERS
    if len(df) >= 4:
        # Use percentiles
        p75 = df[main_metric].quantile(0.75)
        p25 = df[main_metric].quantile(0.25)
        
        results['insights'].append(f"📊 Top 25%: {main_metric} averaged {p75:,.0f} or more")
        results['insights'].append(f"📊 Bottom 25%: {main_metric} averaged {p25:,.0f} or less")
        
        # Create performance bands table
        results['tables']['Performance Bands'] = pd.DataFrame({
            'Band': ['Top 25%', 'Middle 50%', 'Bottom 25%'],
            'Threshold': [f'{p75:,.0f}+', f'{p25:,.0f} - {p75:,.0f}', f'Below {p25:,.0f}']
        })
    
    # 6. SIMPLE AVERAGES
    avg = df[main_metric].mean()
    results['insights'].append(f"📍 Typical value: {main_metric} averages {avg:,.0f}")
    
    # Add summary stats table
    results['tables']['Summary Statistics'] = pd.DataFrame({
        'Metric': numeric_cols,
        'Average': [df[col].mean() for col in numeric_cols],
        'Minimum': [df[col].min() for col in numeric_cols],
        'Maximum': [df[col].max() for col in numeric_cols]
    }).round(0)
    
    # 7. TOTALS
    total = df[main_metric].sum()
    results['insights'].append(f"💯 Total: {main_metric} sums to {total:,.0f}")
    
    return results

def create_executive_summary(insights):
    """Create 3-5 bullet exec summary"""
    summary = []
    
    # Priority order: targets, trends, spikes, segments
    for insight in insights:
        if '✅' in insight or '⚠️' in insight:  # Targets first
            summary.append(insight)
    
    for insight in insights:
        if '📈' in insight or '📉' in insight:  # Trends/spikes
            summary.append(insight)
            if len(summary) >= 3:
                break
    
    # Fill remaining with top insights
    for insight in insights:
        if insight not in summary:
            summary.append(insight)
            if len(summary) >= 5:
                break
    
    # Add action
    if any('⚠️' in s or '📉' in s for s in summary):
        summary.append("💡 Action: Investigate underperforming areas")
    elif any('✅' in s or '📈' in s for s in summary):
        summary.append("💡 Action: Sustain current momentum")
    else:
        summary.append("💡 Action: Continue monitoring trends")
    
    return summary[:5]

def analyze_and_export(input_file, output_file=None):
    """Main function"""
    load_dependencies()

    if output_file is None:
        output_file = Path(input_file).stem + '_analysis.xlsx'
    
    print(f"Reading {input_file}...")
    if input_file.endswith('.csv'):
        df = pd.read_csv(input_file)
    else:
        df = pd.read_excel(input_file)
    
    print(f"Loaded {len(df)} rows, {len(df.columns)} columns")
    
    print("Cleaning data...")
    df = clean_data(df)
    
    print("Analyzing data (7 core analyses)...")
    results = analyze_data(df)
    
    print("Creating summary...")
    exec_summary = create_executive_summary(results['insights'])
    
    print(f"Exporting to {output_file}...")
    
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Tab 1: Summary
        exec_df = pd.DataFrame({'Summary': exec_summary})
        exec_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Tab 2: All Insights
        insights_df = pd.DataFrame({'Insights': results['insights']})
        current_row = 0
        insights_df.to_excel(writer, sheet_name='Full Analysis', index=False, startrow=current_row)
        current_row += len(insights_df) + 3
        
        # Add all analysis tables
        for table_name, table_df in results['tables'].items():
            pd.DataFrame([[table_name]]).to_excel(writer, sheet_name='Full Analysis', 
                                                   startrow=current_row, index=False, header=False)
            current_row += 2
            table_df.to_excel(writer, sheet_name='Full Analysis', startrow=current_row)
            current_row += len(table_df) + 3
        
        # Tab 3: Raw Data
        df.to_excel(writer, sheet_name='Raw Data', index=False)
    
    # Format
    wb = load_workbook(output_file)
    
    # Format Summary
    ws_exec = wb['Summary']
    ws_exec.column_dimensions['A'].width = 100
    ws_exec['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_exec['A1'].font = Font(color='FFFFFF', bold=True, size=12)
    
    for row in range(2, len(exec_summary) + 2):
        ws_exec.cell(row=row, column=1).font = Font(size=11, bold=True)
    
    # Format Full Analysis
    ws_full = wb['Full Analysis']
    ws_full.column_dimensions['A'].width = 80
    ws_full['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_full['A1'].font = Font(color='FFFFFF', bold=True, size=12)
    
    # Format table headers
    for row in ws_full.iter_rows():
        cell = row[0]
        if cell.value and isinstance(cell.value, str):
            if any(keyword in cell.value for keyword in ['Comparison', 'Trend', 'Segment', 'Performance', 'Statistics']):
                cell.font = Font(bold=True, size=12, color='4472C4')
    
    # Format Raw Data
    ws_raw = wb['Raw Data']
    for cell in ws_raw[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    wb.save(output_file)
    
    print(f"\n✅ Analysis complete!")
    print(f"\n📋 SUMMARY:")
    for item in exec_summary:
        print(f"  {item}")
    
    return output_file

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python analyze_data.py <input_file> [output_file]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    analyze_and_export(input_file, output_file)
